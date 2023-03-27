# --------------
# --- IMPORT ---
# --------------

import numpy as np
import pandas as pd
from openpyxl import load_workbook
import glob
from machawai.const import *

# ------------------
# --- EXCEPTIONS ---
# ------------------

class MissingTimeError(Exception):

    def __init__(self, message="No Time values provided") -> None:
        self.message = message
        super().__init__(self.message)

class MissingExtsError(Exception):

    def __init__(self, message="No Extensometer values provided") -> None:
        self.message = message
        super().__init__(self.message)

# ---------------
# --- CLASSES ---
# ---------------

class TestData():
    """
    Holds the detected tensile test data.
    """
    def __init__(self, 
                 disp: 'pd.Series',
                 load: 'pd.Series',
                 exts: 'pd.Series' = None,
                 time: 'pd.Series' = None) -> None:
        """
        `TestData` class constructor.

        Arguments:
        ----------

        disp: Series
            Detected displacement.

        load: Series
            Applied load.

        exts: Series | ndarray
            True specimen extension.

        time: Series | ndarray
            Time data.
        """
        self.disp = self.handleSeriesInput(disp, acceptNone=False, name=DISP)
        self.load = self.handleSeriesInput(load, acceptNone=False, name=LOAD)
        self.exts = self.handleSeriesInput(exts, acceptNone=True, name=EXTS)
        self.time = self.handleSeriesInput(time, acceptNone=True, name=TIME)

        if self.disp.shape[0] != self.load.shape[0]:
            raise ValueError("All specified data series must have the same length.")
        elif isinstance(self.exts, pd.Series) and self.exts.shape[0] != self.load.shape[0]:
            raise ValueError("All specified data series must have the same length.")
        elif isinstance(self.time, pd.Series) and self.time.shape[0] != self.load.shape[0]:
            raise ValueError("All specified data series must have the same length.")

    def handleSeriesInput(self, data: 'pd.Series | np.ndarray | list', acceptNone: bool = False, name: str = "") -> np.ndarray:
        if isinstance(data, pd.Series):
            if data.empty:
                raise ValueError("Error on " + name + " data: empty sequence.")
            data.name = name
            return data
        if isinstance(data, np.ndarray):
            if not data.any():
                raise ValueError("Error on " + name + " data: empty sequence.")
            if len(data.shape) != 1:
                raise ValueError("Error on " + name + " data: invalid shape.")
            return pd.Series(data, name = name)
        if isinstance(data, list):
            if len(data) <= 0:
                raise ValueError("Error on " + name + " data: empty sequence.")
            return pd.Series(data, name = name)
        try:
            if data == None and acceptNone:
                return data
            raise TypeError("Error on " + name + " data: inavlid type.")
        except:
            raise TypeError("Error on " + name + " data: inavlid type.")

    def hasExts(self) -> bool:
        if isinstance(self.exts, pd.Series):
            return not self.exts.empty
        return False
    
    def hasTime(self) -> bool:
        if isinstance(self.time, pd.Series):
            return not self.time.empty
        return False
    
    def getData(self):
        if self.hasTime():
            if self.hasExts():
                return pd.DataFrame({TIME: self.time,
                                     DISP: self.disp,
                                     LOAD: self.load,
                                     EXTS: self.exts})
            return pd.DataFrame({TIME: self.time,
                                 DISP: self.disp,
                                 LOAD: self.load})
        else:
            if self.hasExts():
                return pd.DataFrame({DISP: self.disp,
                                     LOAD: self.load,
                                     EXTS: self.exts})
        return pd.DataFrame({DISP: self.disp,
                             LOAD: self.load})

class SpecimenProperties():
    """
    Holds the geometric properties of the tested specimen.
    """
    def __init__(self,
                 width: 'float | list | np.ndarray',
                 thickness: 'float | list | np.ndarray',
                 interaxis: float,
                 constant_section_length: float,
                 exts_length: float = None) -> None:
        """
        `SpecimenProperties` class constructor.

        Arguments:
        ----------

        width: float | list | np.ndarray
            Mesured specimen width.

        thickness: float | list | np.ndarray
            Mesuerd speciemen thickness.

        interaxis: float
            Interaxis distance.

        exts_length: float
            Initial extensometer length.

        constant_section_length: float
            Constant section length.
        """
        if isinstance(width, float):
            self.width = np.array([width], dtype=float)
        elif isinstance(width, np.ndarray):
            if width.any():
                self.width = width
            else:
                raise ValueError("Empty width.")
        elif isinstance(width, list):
            if len(width) <= 0:
                raise ValueError("Empty width.")
            self.width = np.array(width)
        else:
            raise TypeError("Inavlid data type for width argument.")
        
        if isinstance(thickness, float):
            self.thickness = np.array([thickness], dtype=float)
        elif isinstance(thickness, np.ndarray):
            if thickness.any():
                self.thickness = thickness
            else:
                raise ValueError("Empty thickness.")
        elif isinstance(thickness, list):
            if len(thickness) <= 0:
                raise ValueError("Empty thickness.")
            self.thickness = np.array(thickness)
        else:
            raise TypeError("Inavlid data type for thickness argument.")       

        self.interaxis = interaxis
        self.exts_length = exts_length
        self.constant_section_length = constant_section_length

        self.trasversal_section = self.mean_width() * self.mean_thickness()

    def mean_width(self):
        return self.width.mean()
    
    def mean_thickness(self):
        return self.thickness.mean()
    
    def as_dictionary(self):
        return {
            WIDTH: self.width,
            THICKNESS: self.thickness,
            INTERAXIS: self.interaxis,
            EXTS_LENGTH: self.exts_length,
            CS_LENGTH: self.constant_section_length
        }

class TestSetup():
    """
    Additional Tensile Test information.
    """
    def __init__(self,
                 extensometer_acquired: bool = True,
                 ref_displacement_load: str = MOTOR,
                 ref_param_strain: str = EXTENSOMETER,
                 linearity_dev_method: str = RP02,
                 straingage1_acquired: bool = False,
                 straingage2_acquired: bool = False) -> None:
        """
        `TestSetup` class constructor.

        Arguments:
        ----------

        extensometer_acquired: bool
            `True` if an Extensometer has was used during the Tensile Test.

        ref_displacement_load: int
            Source of Displacement data.

        ref_param_strain: int
            Data source for Strain calculation.

        linearity_dev_method: int
            Linearity Deviation method.

        straingage1_acquired: bool
            TBD

        straingage2_acquired: bool
            TBD
        
        """
        if not ref_displacement_load in ACCEPTED_REF_DISPLACEMENT_LOAD:
            raise ValueError("Invalid ref_displacement_load.")
        if not ref_param_strain in ACCEPTED_REF_PARAM_STRAIN:
            raise ValueError("Invalid ref_param_strain.")
        if not linearity_dev_method in LINEARITY_DEV_METHODS:
            raise ValueError("Invalid linearity_dev_method.")
        
        self.extensometer_acquired = extensometer_acquired
        self.straingage1_acquired = straingage1_acquired
        self.straingage2_acquired = straingage2_acquired
        self.ref_displacement_load = ref_displacement_load
        self.ref_param_strain = ref_param_strain
        self.linearity_dev_method = linearity_dev_method

class CutAndOffset():

    def __init__(self,
                 tail_p: float = 0.0,
                 foot_offset: float = 0.0) -> None:
        """
        Arguments:
        ----------

        tail_p: float:
            Percentage of data points to exclude at the end of the curve.

        foot_offset: float
            Foot offset to apply to data.
        """
        self.tail_p = tail_p
        self.foot_offset = foot_offset

    def apply(self, curve: pd.DataFrame, col_idx: int = 0):
        curve_length = curve.shape[0]
        to_tail = curve_length * self.tail_p
        curve = curve.iloc[:curve_length - to_tail]

        curve.iloc[:,col_idx] = curve.iloc[:,col_idx] - self.foot_offset
        curve = curve.loc[curve.iloc[:,col_idx] < 0]
        return curve

class LinearSection():

    def __init__(self,
                 bottom_cutout: float = 0.0,
                 upper_cutout: float = None) -> None:
        """
        Arguments:
        ----------

        bottom_cutout: float
            Bottom Stress cutout.

        upper_cutout:
            Upper Stress cutout.
        """
        self.bottom_cutout = bottom_cutout
        self.upper_cutout = upper_cutout

    def apply(self, curve: pd.DataFrame, col_idx: int = 1):
        upper_cutout = self.upper_cutout
        if upper_cutout == None:
            upper_cutout = curve.iloc[-1, col_idx]
        curve = curve.loc[curve.iloc[:,col_idx] >= self.bottom_cutout]
        curve = curve.loc[curve.iloc[:,col_idx] <= upper_cutout]

        return curve

class TensileTest():
    """
    Describe a Tensile Test.
    """
    def __init__(self, 
                 testData: TestData,
                 specimenProperties: SpecimenProperties,
                 testSetup: TestSetup,
                 cutAndOffset: CutAndOffset = None,
                 linearSection: LinearSection = None,
                 filename: str = None) -> None:
        """
        `TensileTest` class constructor.

        Arguments:
        ----------

        testData: `TestData`
            Detected tensile test data.

        specimenProperties: `SpecimenProperties`
            Geometric properties of the tested specimen.

        testSetup: `TestSetup`
            Additional test info.

        cutAndOffset: `CutAndOffset`
            Cut and offset options.

        linearSection: `LinearSection`
            Options for linear section localization.

        filename: str:
            Name of source file.
        """
        self.testData = testData
        self.specimenProperties = specimenProperties
        self.testSetup = testSetup
        self.cutAndOffset = cutAndOffset
        self.linearSection = linearSection
        self.filename = filename

        if self.linearSection == None:
            self.linearSection = LinearSection()

        if self.testSetup.extensometer_acquired and not self.testData.hasExts():
            raise ValueError("Inconsistent setup and data: Extensometer is acquired but no value has been provided.")

    def getTime(self):
        """
        Return the Time values.
        """
        if self.testData.hasTime():
            return self.testData.time
        raise MissingTimeError()
    
    def getDisplacement(self):
        """
        Return the Displacement values.
        """
        return self.testData.disp
    
    def getLoad(self):
        """
        Return the Load values.
        """
        return self.testData.load
    
    def getExtensometer(self):
        """
        Return the Extensometer values.
        """
        if self.testData.hasExts():
            return self.testData.exts
        raise MissingExtsError()
    
    def getDispStrain(self):
        """
        Compute the Strain values on Displacement.
        """
        return self.getDisplacement() / self.specimenProperties.exts_length
    
    def getExtsStrain(self):
        """
        Compute the Strain values on Extensometer.
        """
        if self.testData.hasExts():
            strain = self.getExtensometer() / self.specimenProperties.exts_length
            strain.name = EXTS_STRAIN
            return strain
        else:
            raise MissingExtsError()
    
    def getStress(self):
        """
        Compute the Stress values.
        """
        stress = self.getLoad() / self.specimenProperties.trasversal_section
        stress.name = STRESS
        return stress
    
    def getData(self):
        """
        Return the Test Data as a single Dataframe.
        """
        data = self.testData.getData()
        # Add disp/extx strain
        data[DISP_STRAIN] = self.getDispStrain()
        if self.testData.hasExts():
            data[EXTS_STRAIN] = self.getExtsStrain()
        data[STRESS] = self.getStress()

        return data
    
    def getDataLabels(self):
        """
        Return test Data labels.
        """
        return self.getData().columns
    
    def selectData(self, labels: 'list[str]'):
        """
        Filter Test Data given a label list.
        """
        return self.getData()[labels]
    
    def getTimeAtIndex(self, idx: int):
        """
        Return the Time value at the given index.
        """
        if self.testData.hasTime():
            return self.testData.time[idx]
        raise MissingTimeError()
    
    def getDispAtIndex(self, idx: int):
        """
        Return the Displacement value at the given index.
        """
        return self.testData.disp[idx]

    def getLoadAtIndex(self, idx: int):
        """
        Return the Load value at the given index.
        """
        return self.testData.load[idx]
    
    def getExtsAtIndex(self, idx: int):
        """
        Return the Extensometer value at the given index.
        """
        if self.testData.hasExts():
            return self.testData.exts[idx]
        raise MissingExtsError()

    def getStrainAtIndex(self, idx: int, use_displacement: bool = False):
        """
        Return the Strain value at the given index.
        """
        if use_displacement:
            return self.getDispStrain()[idx]
        else:
            return self.getExtsStrain()[idx]

    def getMaxLoad(self):
        """
        Return the maximum Load value and its index.
        """
        max_load_idx = self.testData.load.argmax()
        return self.testData.load[max_load_idx], max_load_idx
    
    def getMaxStress(self):
        """
        Return the maximum Stress value and its index.
        """
        max_load, max_load_idx = self.getMaxLoad()
        return max_load / self.specimenProperties.trasversal_section, max_load_idx

    def getLoadDisplacementCurve(self, cut_and_offset: bool = True):
        """
        Return the Load-Displacement curve.
        """
        curve = pd.DataFrame({DISP: self.testData.disp,
                              LOAD: self.testData.load})
        if cut_and_offset and self.cutAndOffset != None:
            curve = self.cutAndOffset.apply(curve=curve, col_idx=0)
        return curve
    
    def getLoadExtsCurve(self, cut_and_offset: bool = True):
        """
        Return the Load-Extensometer curve.

        Raise `MissingExtsError` if no Extensometer value is provided.
        """
        if not self.testData.hasExts():
            raise MissingExtsError()
        curve = pd.DataFrame({EXTS: self.testData.exts,
                              LOAD: self.testData.load})
        if cut_and_offset and self.cutAndOffset != None:
            curve = self.cutAndOffset.apply(curve=curve, col_idx=0)
        return curve
    
    def getStrainStressCurve(self, use_displacement: bool = False, cut_and_offset: bool = True):
        """
        Return the Strain-Stress curve.

        Arguments:
        ----------

        use_displacement: bool
            If `True` displacement values are used for Strain calculation.
        """
        stress = self.getStress()
        if use_displacement:
            strain = self.getDispStrain()
            curve = pd.DataFrame({DISP_STRAIN: strain,
                                  STRESS: stress})
        else:
            strain = self.getExtsStrain()
            curve = pd.DataFrame({EXTS_STRAIN: strain,
                                  STRESS: stress})
        if cut_and_offset and self.cutAndOffset != None:
            curve = self.cutAndOffset.apply(curve=curve, col_idx=0)
        return curve
        
    def getProperties(self):
        """ 
        Return the Speciemn Properties as a dictionary 
        """
        return self.specimenProperties.as_dictionary()
        
    def selectProperties(self, labels: 'list[str]'):
        """
        Filter Specimen Properties given a label list.
        """
        select = {}
        sporp = self.specimenProperties.as_dictionary()
        for label in labels:
            select[label] = sporp[label]
        return select
    
    def getLinearSection(self, use_displacement: bool = False):
        if use_displacement:
            curve = self.getLoadDisplacementCurve()
        else:
            curve = self.getLoadExtsCurve() 
        return self.linearSection.apply(curve=curve)
    
    def getDataStats(self, labels: 'list[str]' = None):
        if labels == None:
            return self.getData().describe()
        else:
            return self.selectData(labels=labels).describe()

# -----------------
# --- FUNCTIONS ---
# -----------------

def string2bool(string: 'str | int') -> bool:
    if isinstance(string, int):
        return string == 1
    return string.lower() in ["true", "t", "1", "yes", "y", "s"]

def readTensileTest(file: str) -> TensileTest:
    """
    Read a Tensile Test from file.
    """
    # Load the workbook
    wb = load_workbook(filename=file, data_only=True)
    # Get the sheet names
    sheet_names = wb.sheetnames
    # Check that RAW end ELAB sheets exist
    assert RAW in sheet_names and ELAB in sheet_names
    # 1) Read the Test data
    df = pd.read_excel(file, sheet_name=RAW, header=[0,1], dtype=float)
    testData = TestData(disp = df[DISP].iloc[:,0],
                        load = df[LOAD].iloc[:,0],
                        exts = df[EXTS].iloc[:,0] if EXTS in df.columns else None,
                        time = df[TIME].iloc[:,0] if TIME in df.columns else None)
    
    # Read Specimen Properties and options
    prop_sheet = wb[ELAB]
    # Read properties table
    table = prop_sheet[SPROP_START: SPROP_END]

    # 2) Specimen Properties
    width = table[0]
    thickness = table[1]
    interaxis = table[2][0].value
    constant_section_length = table[3][0].value
    exts_length = table[4][0].value
    # Width and Thickness needs to be converted from tuple to list
    width = list(map(lambda c: c.value, width))
    thickness = list(map(lambda c: c.value, thickness))
    sprop = SpecimenProperties(width=width,
                               thickness=thickness,
                               interaxis=interaxis,
                               constant_section_length=constant_section_length,
                               exts_length=exts_length)
    
    # 3) Test Setup
    table = prop_sheet[SETUP_START: SETUP_END]
    ext_acquired = string2bool(table[0][0].value)
    s1_acquired = string2bool(table[1][0].value)
    s2_acquired = string2bool(table[2][0].value)
    rdl = table[3][0].value
    if not (rdl in ACCEPTED_REF_DISPLACEMENT_LOAD):
        raise ValueError("Unknown Ref. Displacement Load. Given value: {}. Accepted: {}".format(rdl, ACCEPTED_REF_DISPLACEMENT_LOAD))
    rps = table[4][0].value
    if not (rps in ACCEPTED_REF_PARAM_STRAIN):
        raise ValueError("Unknown Ref. Parameter for Strain Calculation. Given value: {}. Accepted: {}".format(rps, ACCEPTED_REF_PARAM_STRAIN))
    ldm = table[5][0].value
    if not (ldm in LINEARITY_DEV_METHODS):
        raise ValueError("Unknow Linearity Deviation Method. Given: {}. Accepted: {}".format(ldm, LINEARITY_DEV_METHODS))
    setup = TestSetup(extensometer_acquired = ext_acquired,
                      ref_displacement_load = rdl,
                      ref_param_strain = rps,
                      linearity_dev_method = ldm,
                      straingage1_acquired = s1_acquired,
                      straingage2_acquired = s2_acquired)
    
    # 4) Cut and Offset
    tp = prop_sheet[TAIL_P].value / 100
    foot_offset = prop_sheet[FOOT_OFFSET].value / 100
    cao = CutAndOffset(tail_p=tp, foot_offset=foot_offset)

    # 5) Linear Section
    bc = prop_sheet[BOTTOM_CUTOUT]
    uc = prop_sheet[UPPER_CUTOUT]
    linsec = LinearSection(bottom_cutout=bc, upper_cutout=uc)

    filename = file[file.rindex("\\") + 1:]

    return TensileTest(testData=testData,
                       specimenProperties=sprop,
                       testSetup=setup,
                       cutAndOffset=cao,
                       linearSection=linsec,
                       filename=filename)

def readTensileTestCollection(root: 'str | list[str]', exts: 'list[str]' = ['xlsx']) -> 'list[TensileTest]':
    # TODO: ottimizzare gestione estensioni
    """
    Read a collection of Tensile Tests from a root directory.
    """
    tt_collection = []
    if isinstance(root, str):
        root = [root]
    for dir in root:
        filenames = []
        for ext in exts:
            filenames.extend(glob.glob(dir + "/*." + ext))
        for file in filenames:
            try:
                tt_collection.append(readTensileTest(file=file))
            except:
                print("An error occured while reading file {}".format(file))
    return tt_collection
